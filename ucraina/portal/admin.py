from django.contrib import admin
from django.http import HttpResponse
from openpyxl.writer.excel import save_virtual_workbook
from .models import Richiesta, Componente, UfficioCompetente, Prestazione, StatoPrestazione, ListaPrestazioni
from django_object_actions import DjangoObjectActions
from openpyxl import Workbook
from openpyxl.worksheet.table import Table
from openpyxl.utils import get_column_letter
from daterangefilter.filters import PastDateRangeFilter
from django.utils.html import format_html
import datetime

# Register your models here.
admin.site.register(Componente)
admin.site.register(UfficioCompetente)
admin.site.register(Prestazione)
admin.site.register(StatoPrestazione)
admin.site.register(ListaPrestazioni)


class ComponenteInline(admin.StackedInline):
    model = Componente
    fields = [('cognome', 'nome'), ('luogo_nascita', 'data_nascita'), 'note']
    extra = 0


class PrestazioneInline(admin.StackedInline):
    model = Prestazione
    fields = ['descrizione', 'nota', 'stato']
    extra = 0
    max_num = 5


class RichiestaAdmin(DjangoObjectActions, admin.ModelAdmin):
    list_display = (
        'get_componenti',
        #'id',
        'data_colloquio',
        'telefono_1',
        #'telefono_1_note',
        'telefono_2',
        #'telefono_2_note',
        'email',
        'collocazione_attuale',
        'ufficio_competente',
        'visualizza'
    )

    search_fields = (
        'componente__nome',
        'componente__cognome',
        'id',
        'data_colloquio',
        'telefono_1',
        'telefono_2',
        'email',
        'ufficio_competente__nome'
    )

    fields = [
        ('id', 'data_colloquio'),
        ('telefono_1', 'telefono_1_note'),
        ('telefono_2', 'telefono_2_note'),
        'email',
        'collocazione_attuale',
        'ufficio_competente'
    ]

    inlines = [ComponenteInline, PrestazioneInline]

    list_filter = [
        ('data_colloquio', PastDateRangeFilter),
        'ufficio_competente'
    ]

    def export_excel(modeladmin, request, queryset):
        print("PREMUTO EXPORT")
        """Esportare dati in formato Excel"""
        report = ''
        wb = Workbook()
        sheet = wb.active
        row = 1

        sheet.cell(row=row, column=1).value = "DATA COLLOQUIO"
        sheet.cell(row=row, column=2).value = "ID RICHIESTA"
        sheet.cell(row=row, column=3).value = "COMPONENTI NUCLEO"
        sheet.cell(row=row, column=4).value = "TELEFONO 1"
        sheet.cell(row=row, column=5).value = "TELEFONO 1 NOTE"
        sheet.cell(row=row, column=6).value = "TELEFONO 2"
        sheet.cell(row=row, column=7).value = "TELEFONO 2 NOTE"
        sheet.cell(row=row, column=8).value = "EMAIL"
        sheet.cell(row=row, column=9).value = "COLLOCAZIONE ATTUALE"
        sheet.cell(row=row, column=10).value = "UFFICIO COMPETENTE"
        sheet.cell(row=row, column=11).value = "PRESTAZIONE 1"
        sheet.cell(row=row, column=12).value = "STATO PREST 1"
        sheet.cell(row=row, column=13).value = "PRESTAZIONE 2"
        sheet.cell(row=row, column=14).value = "STATO PREST 2"
        sheet.cell(row=row, column=15).value = "PRESTAZIONE 3"
        sheet.cell(row=row, column=16).value = "STATO PREST 3"
        sheet.cell(row=row, column=17).value = "PRESTAZIONE 4"
        sheet.cell(row=row, column=18).value = "STATO PREST 4"
        sheet.cell(row=row, column=19).value = "PRESTAZIONE 5"
        sheet.cell(row=row, column=20).value = "STATO PREST 5"
        richieste = Richiesta.objects.all()
        for rich in richieste:
            row = row + 1
            print("RICHIESTA")
            print(rich)
            sheet.cell(row=row, column=2).value = str(rich)
            id_rich = rich.id
            sheet.cell(row=row, column=1).value = str(rich.data_colloquio)
            data_rich = rich.data_colloquio
            sheet.cell(row=row, column=4).value = str(rich.telefono_1)
            tel1_rich = rich.telefono_1
            sheet.cell(row=row, column=5).value = str(rich.telefono_1_note)
            tel1n_rich = rich.telefono_1_note
            sheet.cell(row=row, column=6).value = str(rich.telefono_2)
            tel2_rich = rich.telefono_2
            sheet.cell(row=row, column=7).value = str(rich.telefono_2_note)
            tel2n_rich = rich.telefono_2_note
            sheet.cell(row=row, column=9).value = str(rich.collocazione_attuale)
            coll_rich = rich.collocazione_attuale
            sheet.cell(row=row, column=10).value = str(rich.ufficio_competente)
            uffc_rich = rich.ufficio_competente
            sheet.cell(row=row, column=8).value = str(rich.email)
            email_rich = rich.email
            lista_comp = ''
            componenti = Componente.objects.filter(richiesta_id=id_rich)
            for comp in componenti:
                lista_comp += comp.cognome + ' ' + comp.nome + ', '
            lista_comp = lista_comp[:-2]
            sheet.cell(row=row, column=3).value = lista_comp
            print("COMPONENTI")
            print(lista_comp)
            lista_prest = ''
            prestazioni = Prestazione.objects.filter(richiesta_id=id_rich)
            k = 0
            for prest in prestazioni:
                lista_prest += str(prest.descrizione) + ' (' + str(prest.stato.descrizione) + '), '
                print(lista_prest)
                sheet.cell(row=row, column=11 + k).value = str(prest.descrizione)
                sheet.cell(row=row, column=12 + k).value = str(prest.stato.descrizione)
                k = k + 2
            print("PRESTAZIONI")
            print(lista_prest)
            print("--------")

        table = Table(displayName="Table1", ref="A1:" + get_column_letter(sheet.max_column) + str(sheet.max_row))

        sheet.add_table(table)

        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/ms-excel', )
        response['Content-Disposition'] = f'attachment; filename={datetime.datetime.now().strftime("%Y%m%d%H%M")}-Export_excel.xlsx'

        return response

    changelist_actions = ("export_excel",)

    def get_componenti(self, instance):
        componenti = Componente.objects.filter(richiesta__id=instance.id)
        result = ''
        for comp in componenti:
            result += comp.cognome + ' ' + comp.nome + ', '
        return result[:-2]

    get_componenti.short_description = 'Nucleo familiare'

    def visualizza(self, obj):
        return format_html(f'<a href="/admin/portal/richiesta/{obj.id}/change/" class==default">Visualizza</a>')

admin.site.register(Richiesta, RichiestaAdmin)
